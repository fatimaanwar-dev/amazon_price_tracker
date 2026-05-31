# ================================================
# Amazon Price Tracker v2
# Author: fafay | Fiverr.com/fafay
# Description: Tracks Amazon product prices and
#              saves history to an Excel file.
#              Uses a real Chromium browser to
#              avoid Amazon bot detection.
# ================================================

from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import time
import random
import re

EXCEL_FILE = "price_history.xlsx"


def normalize_url(url: str) -> str:
    """Clean up any Amazon URL format into a standard /dp/ASIN link."""
    url = url.strip()
    asin_match = re.search(r"/(?:dp|gp/product|ASIN)/([A-Z0-9]{10})", url)
    if asin_match:
        return f"https://www.amazon.com/dp/{asin_match.group(1)}"
    return url


def get_discount_info(soup) -> dict:
    """Extract discount percentage and original price from product page."""
    discount_info = {"discount_percent": None, "original_price": None}
    
    try:
        # Look for discount percentage badge
        discount_span = soup.find("span", {"class": re.compile(r".*?savingsPercent.*?")})
        if discount_span:
            discount_text = discount_span.get_text(strip=True)
            percent_match = re.search(r"(\d+)%", discount_text)
            if percent_match:
                discount_info["discount_percent"] = f"{percent_match.group(1)}%"
        
        # Look for original price (usually with strikethrough)
        price_wrapper = soup.find("span", {"class": re.compile(r".*?priceBlockStrikeprice.*?")})
        if price_wrapper:
            price_text = price_wrapper.get_text(strip=True)
            price_match = re.search(r"[\$£€][\s]?([\d.]+)", price_text)
            if price_match:
                discount_info["original_price"] = float(price_match.group(1))
        
        # Alternative: look for list price in product information
        if not discount_info["original_price"]:
            list_price = soup.find("span", {"class": re.compile(r".*?listPrice.*?")})
            if list_price:
                price_text = list_price.get_text(strip=True)
                price_match = re.search(r"[\$£€][\s]?([\d.]+)", price_text)
                if price_match:
                    discount_info["original_price"] = float(price_match.group(1))
    
    except Exception as e:
        pass
    
    return discount_info


def get_amazon_price(url: str, page) -> dict:
    """Scrape product name, price, and discount info using a real browser page."""
    url = normalize_url(url)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    try:
        # Random human-like delay
        time.sleep(random.uniform(2, 5))

        page.goto(url, timeout=20000, wait_until="domcontentloaded")
        time.sleep(random.uniform(2, 3))  # wait for JS to load prices

        content = page.content()
        soup = BeautifulSoup(content, "html.parser")

        # CAPTCHA / bot check detection
        if "Type the characters" in content or "robot check" in content.lower():
            print("  [!] Amazon CAPTCHA detected — wait a few minutes and retry.")
            return {"title": "CAPTCHA Block", "price": None, "discount_percent": None, 
                    "original_price": None, "url": url, "timestamp": timestamp}

        # Product title
        title_tag = soup.find(id="productTitle")
        title = title_tag.get_text(strip=True) if title_tag else None
        if not title:
            og = soup.find("meta", property="og:title")
            title = og["content"].strip() if og else "Unknown Product"
        title = title[:60]

        # Price — try multiple selectors
        price = None
        for selector in ["#priceblock_ourprice", "#priceblock_dealprice",
                         ".a-price .a-offscreen", ".a-price-whole"]:
            el = soup.select_one(selector)
            if el:
                raw = re.sub(r"[^\d.]", "", el.get_text(strip=True))
                try:
                    price = float(raw)
                    break
                except Exception:
                    continue

        # Get discount information
        discount_data = get_discount_info(soup)

        return {"title": title, "price": price, "discount_percent": discount_data["discount_percent"],
                "original_price": discount_data["original_price"], "url": url, "timestamp": timestamp}

    except Exception as e:
        print(f"  [!] Error: {e}")
        return {"title": "Error", "price": None, "discount_percent": None, 
                "original_price": None, "url": url, "timestamp": timestamp}


def style_header_row(ws, row, num_cols):
    fill = PatternFill("solid", start_color="1F4E79")
    font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align


def style_data_row(ws, row, num_cols, alternate):
    fill = PatternFill("solid", start_color="D6E4F0" if alternate else "FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.alignment = align
        c.border = border
        c.font = Font(name="Arial", size=10)


def setup_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Price History"
    headers = ["#", "Product Name", "Price (€/$)", "Original Price", "Discount %", "Change", "Date & Time", "URL"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22
    for i, w in enumerate([5, 45, 14, 16, 14, 16, 20, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb


def load_or_create_workbook():
    """Always create a fresh workbook with correct structure."""
    return setup_workbook()


def get_last_price(ws, url):
    """Find most recent price for a URL — matches on ASIN so format doesn't matter."""
    asin_match = re.search(r"/dp/([A-Z0-9]{10})", url)
    asin = asin_match.group(1) if asin_match else url

    last = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        stored_url = str(row[7]) if row[7] else ""  # URL is now in column 8 (index 7)
        if asin in stored_url and row[2] not in (None, "N/A"):
            try:
                last = float(row[2])
            except Exception:
                pass
    return last


def append_record(ws, counter, data, last_price):
    price = data["price"]
    change = ""
    if price and last_price:
        diff = round(price - last_price, 2)
        change = f"▲ +{diff}" if diff > 0 else (f"▼ {diff}" if diff < 0 else "→ No change")

    discount_display = data.get("discount_percent") or "N/A"
    original_price_display = data.get("original_price") or "N/A"
    
    row_data = [counter, data["title"], price or "N/A", original_price_display, 
                discount_display, change, data["timestamp"], data["url"]]
    ws.append(row_data)
    rn = ws.max_row
    style_data_row(ws, rn, len(row_data), rn % 2 == 0)

    if change:
        cc = ws.cell(row=rn, column=6)
        cc.font = Font(name="Arial", size=10, bold=True,
                       color="C00000" if "▲" in change else "375623")


def track_prices(urls):
    print("\n🔍 Amazon Price Tracker v2 — fafay | Fiverr\n" + "=" * 45)
    wb = load_or_create_workbook()
    ws = wb.active
    counter = ws.max_row  # row 1 = header, so counter starts at correct next number

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800},
            locale="en-US",
        )
        page = context.new_page()

        for url in urls:
            print(f"\n📦 Checking: {url[:70]}...")
            data = get_amazon_price(url, page)
            last_price = get_last_price(ws, data["url"])
            append_record(ws, counter, data, last_price)
            counter += 1
            discount_str = f" | Discount: {data['discount_percent']}" if data.get('discount_percent') else ""
            original_str = f" | Original: {data['original_price']}" if data.get('original_price') else ""
            print(f"  ✅ {data['title']} — Price: {data['price']}{original_str}{discount_str}")

        browser.close()

    safe_save_workbook(wb, EXCEL_FILE)


def safe_save_workbook(wb, filename, max_retries=3):
    """Try to save workbook, retry if file is locked (e.g., open in Excel)."""
    for attempt in range(max_retries):
        try:
            wb.save(filename)
            print(f"\n✅ Done! Saved to '{filename}'\n")
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"  ⚠️  File is locked (possibly open in Excel). Retrying in 2 seconds...")
                time.sleep(2)
            else:
                print(f"  ❌ Could not save '{filename}' — file is open in Excel or another app.")
                print(f"  💡 Close the file and run the script again.")
                return False
    return False


# ── ADD YOUR AMAZON PRODUCT URLs HERE ─────────────
PRODUCT_URLS = [
    "https://a.co/d/09mVL2DA",   # Echo Dot — standard URL
    "https://www.amazon.com/Sampeel-Palazzo-Trousers-Vacation-Clothing/dp/B0GHY8ZV8S?pd_rd_w=nyODL&content-id=amzn1.sym.7154b216-2f41-4765-8779-9a282bdc557c&pf_rd_p=7154b216-2f41-4765-8779-9a282bdc557c&pf_rd_r=64HV5RPYT58E5DJ3J2S4&pd_rd_wg=0RNmU&pd_rd_r=d9d99d88-5cfb-4d3c-82cf-eac2d486ed37&ref_=sspa_dk_detail_sbb_img_0&sp_csd=d2lkZ2V0TmFtZT1zcF9kZXRhaWxfdGhlbWF0aWM&th=1&psc=1",   # Fire TV Stick
    "https://a.co/d/03caNzHd","https://www.amazon.com/SUPFINE-Compatible-Protection-Translucent-Shockproof/dp/B0DQKVJ2TL?ref=dlx_deals_dg_dcl_B0DQKVJ2TL_dt_sl14_f1_pi&pf_rd_r=M23RQAA66XVBAJTJCNRZ&pf_rd_p=1ca387c6-cdd2-47fc-9c79-443ca8c7d4f1&sbo=RZvfv%2F%2FHxDF%2BO5021pAnSA%3D%3D&th=1"
    # Works with: short links, long URLs, /gp/product/ links — paste any Amazon URL!
]

if __name__ == "__main__":
    track_prices(PRODUCT_URLS)
